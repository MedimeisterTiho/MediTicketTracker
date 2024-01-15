from selenium import webdriver
import time
import pandas as pd
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
import xlsxwriter
from xlsxwriter import Workbook
import openpyxl


Service_object = Service('./chromedriver')

#lädt Codes
Filexlx = openpyxl.load_workbook('Ticketcodes.xlsx')
sheet = Filexlx['Daten']
Codesnicht = Filexlx['Freie Codes']

print('Medimeisterschaften 2024')


#Lädt Chrome

driver = webdriver.Chrome()
driver.get("https://ticket.medimeisterschaften.com/")
exp_url = "https://ticket.medimeisterschaften.com/?voucher_invalid"


i = 1
t = 2

time.sleep(2)

#Definiert die Suchleiste und den Button
search_bar = driver.find_element(By.ID,'voucher')
button = driver.find_element(By.XPATH, "/html/body/div[1]/aside[1]/form/div/div[2]/button")

#Leert Tabelle
for row in Codesnicht['A2:F2000']:
        for cell in row:
            cell.value = None

#Überprüft die Codes
try:
    while  i<4000:
        time.sleep(0.5)
        if sheet.cell(row=i+1,column=6).value == 'eingelöst':
             i=i+1
        else:
            search_bar.clear()
            Ticket = sheet.cell(row=i+1, column=5).value
            search_bar.send_keys(Ticket)
        
            button.click()
        
            cur_url = driver.current_url
            if exp_url == cur_url:
            
                driver.back()
                
                sheet.cell(row=i+1, column=6).value = 'eingelöst'
                print(sheet.cell(row=i+1, column=5).value + ' eingelöst')       
            else:
            
                driver.back()
                
                sheet.cell(row=i+1, column=6).value = 'nicht eingelöst'
                Codesnicht.cell(row=t, column=6).value = 'nicht eingelöst'
                print(sheet.cell(row=i+1, column=5).value + ' nicht eingelöst')  
                Codesnicht.cell(row=t, column=1).value = sheet.cell(row=i+1,column=1).value
                Codesnicht.cell(row=t, column=2).value = sheet.cell(row=i+1,column=2).value
                Codesnicht.cell(row=t, column=3).value = sheet.cell(row=i+1,column=3).value
                Codesnicht.cell(row=t, column=4).value = sheet.cell(row=i+1,column=4).value
                Codesnicht.cell(row=t, column=5).value = sheet.cell(row=i+1,column=5).value
                t = t + 1 
            i=i+1      
finally:
    driver.close()
    Filexlx.save('Ticketcodes.xlsx')
